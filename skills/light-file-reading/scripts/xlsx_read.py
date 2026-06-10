#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX 读取与数据画像（light-file-reading）。

依赖：openpyxl（读公式/格式/多 sheet），pandas（数据画像）。
铁律：openpyxl 无求值引擎——公式只存字符串。要算值用 data_only=True 读缓存，
或写后跑 LibreOffice 重算（见 references/XLSX-REF.md）。
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")


def list_sheets(path):
    """返回 sheet 名列表。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_formulas(path, sheet=None):
    """读单元格里的公式（不求值）→ [(coord, formula)]。
    sheet=None 取活动表。公式以 '=' 开头。"""
    from openpyxl import load_workbook
    wb = load_workbook(path)  # data_only=False，保留公式字符串
    ws = wb[sheet] if sheet else wb.active
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out.append((cell.coordinate, cell.value))
    wb.close()
    return out


def read_values(path, sheet=None):
    """读缓存计算值（data_only=True）→ [[...]]。
    注意：需 Excel/LibreOffice 之前算过并保存，否则公式格缓存为 None。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    return rows


def profile(path, sheet=0):
    """用 pandas 做数据画像：返回 {shape, columns, dtypes, describe}。
    sheet 可为索引或名字。"""
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet)
    return {
        "shape": df.shape,
        "columns": list(df.columns),
        "dtypes": {c: str(t) for c, t in df.dtypes.items()},
        "describe": df.describe(include="all").to_dict(),
    }


def _selftest():
    """合成 xlsx（含公式）跑全流程，断言后清理。"""
    import os
    import tempfile
    from openpyxl import Workbook

    tmp = tempfile.mkdtemp(prefix="xlsxread_")
    src = os.path.join(tmp, "sample.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["n", "sq"])
    for i in range(1, 5):
        ws.append([i, i * i])
    ws["D1"] = "=SUM(A2:A5)"  # 公式格
    wb.create_sheet("Empty")
    wb.save(src)

    sheets = list_sheets(src)
    assert sheets == ["Data", "Empty"], sheets

    formulas = read_formulas(src, "Data")
    assert ("D1", "=SUM(A2:A5)") in formulas, formulas

    vals = read_values(src, "Data")
    assert vals[0][0] == "n" and vals[1][:2] == [1, 1], vals[:2]

    prof = profile(src, sheet="Data")
    assert prof["shape"] == (4, 2) and prof["columns"] == ["n", "sq"], prof

    import shutil
    shutil.rmtree(tmp, ignore_errors=True)
    print("xlsx_read self-test OK: sheets/formulas/values/profile all passed")


if __name__ == "__main__":
    if len(sys.argv) > 2 or (len(sys.argv) == 2 and sys.argv[1] != "--selftest"):
        raise SystemExit(f"usage: python {sys.argv[0]} [--selftest]")
    _selftest()
